"""
額外部分 — 將資料庫裁判書資料匯出為 Excel 檔案

使用方式:
    python export_excel.py                          # 匯出最新 100 筆
    python export_excel.py -n 50                    # 匯出最新 50 筆
    python export_excel.py -n 0                     # 匯出全部
    python export_excel.py -n 30 -k 詐欺            # 篩選關鍵字
    python export_excel.py -n 30 -c 臺灣臺北        # 篩選法院
    python export_excel.py -n 30 --full-text        # 含全文欄位
    python export_excel.py -n 20 -o my_export.xlsx  # 指定輸出檔名
"""

import sys
import io
import sqlite3
import argparse
import logging
from datetime import datetime
from pathlib import Path

# Windows terminal 預設 cp1252，強制 UTF-8 避免中文印出錯誤
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from typing import Optional, List, Dict
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ─── 設定 ────────────────────────────────────────────────────────────────────
DB_PATH = "judgments.db"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger(__name__)

# 資料欄位對應（DB 欄位名 → 顯示名稱）
BASE_COLUMNS: List[tuple] = [
    ("case_number",       "裁判字號"),
    ("source_url",        "裁判書連結"),
    ("court",             "法院"),
    ("judgment_date",     "裁判日期"),
    ("case_type",         "案件類型"),
    ("judgment_type",     "裁判種類"),
    ("party_roles",       "當事人角色"),
    ("plaintiff",         "原告／聲請人"),
    ("plaintiff_agent",   "原告代理人"),
    ("defendant",         "被告／相對人"),
    ("defendant_agent",   "被告代理人"),
    ("appellant",         "上訴人"),
    ("appellant_agent",   "上訴人代理人"),
    ("appellee",          "被上訴人"),
    ("appellee_agent",    "被上訴人代理人"),
    ("verdict",           "主文"),
    ("facts",             "事實"),
    ("facts_and_reasons", "事實及理由"),
    ("criminal_facts",    "犯罪事實"),
    ("reasons",           "理由"),
    ("conclusion",        "結論"),
    ("applicable_laws",   "適用法條"),
    ("judges",            "法官"),
    ("clerk",             "書記官"),
    ("keyword",           "搜尋關鍵字"),
    ("parsed_at",         "解析時間"),
]

# 每欄的建議寬度（字元數）
COL_WIDTHS: Dict[str, int] = {
    "序號": 6, "裁判字號": 26, "裁判書連結": 55, "法院": 22, "裁判日期": 16,
    "案件類型": 22, "裁判種類": 10, "當事人角色": 40,
    "原告／聲請人": 22, "原告代理人": 22, "被告／相對人": 22, "被告代理人": 22,
    "上訴人": 22, "上訴人代理人": 22, "被上訴人": 22, "被上訴人代理人": 22,
    "主文": 45, "事實": 55, "事實及理由": 65, "犯罪事實": 55,
    "理由": 65, "結論": 35, "適用法條": 35,
    "法官": 28, "書記官": 16, "搜尋關鍵字": 16, "解析時間": 22, "全文": 65,
}


# ─── 資料擷取 ─────────────────────────────────────────────────────────────────
def fetch_judgments(
    limit:      Optional[int] = 100,
    offset:     int = 0,
    keyword:    Optional[str] = None,
    court:      Optional[str] = None,
    start_date: Optional[str] = None,
    end_date:   Optional[str] = None,
) -> List[Dict]:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    sql    = "SELECT * FROM judgments WHERE 1=1"
    params: list = []

    if keyword:
        sql += " AND (keyword LIKE ? OR case_number LIKE ? OR verdict LIKE ?)"
        p    = f"%{keyword}%"
        params.extend([p, p, p])
    if court:
        sql += " AND court LIKE ?"
        params.append(f"%{court}%")
    if start_date:
        sql += " AND judgment_date >= ?"
        params.append(start_date)
    if end_date:
        sql += " AND judgment_date <= ?"
        params.append(end_date)

    sql += " ORDER BY parsed_at DESC"
    if limit:
        sql += f" LIMIT {limit} OFFSET {offset}"

    rows = [dict(r) for r in conn.execute(sql, params).fetchall()]
    conn.close()
    logger.info("Fetched %d records from DB", len(rows))
    return rows


# ─── Excel 格式化輔助 ─────────────────────────────────────────────────────────
def _header_style(ws) -> None:
    fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    font = Font(color="FFFFFF", bold=True, size=11, name="微軟正黑體")
    aln  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = aln
    ws.row_dimensions[1].height = 32


def _data_style(ws, n_rows: int) -> None:
    even_fill = PatternFill(start_color="EEF2F7", end_color="EEF2F7", fill_type="solid")
    aln        = Alignment(vertical="top", wrap_text=True)
    for row_idx in range(2, n_rows + 2):
        ws.row_dimensions[row_idx].height = 70
        for cell in ws[row_idx]:
            cell.alignment = aln
            if row_idx % 2 == 0:
                cell.fill = even_fill


def _set_col_widths(ws, col_names: List[str]) -> None:
    for col_idx, name in enumerate(col_names, 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = COL_WIDTHS.get(name, 20)


# ─── 核心匯出函式 ─────────────────────────────────────────────────────────────
def export_to_excel(
    rows:             List[Dict],
    output_path:      str,
    include_full_text: bool = False,
) -> bool:
    if not rows:
        logger.warning("No data to export.")
        return False

    columns = list(BASE_COLUMNS)
    if include_full_text:
        columns.append(("full_text", "全文"))

    # ── 建立 DataFrame ─────────────────────────────────────────────
    records = []
    for row in rows:
        entry: dict = {}
        for field, label in columns:
            val = str(row.get(field, "") or "")
            # Excel 單格上限 32767 字
            if len(val) > 32700:
                val = val[:32700] + "…(截斷)"
            entry[label] = val
        records.append(entry)

    col_labels = [label for _, label in columns]
    df = pd.DataFrame(records, columns=col_labels)

    # ── 寫入 Excel ─────────────────────────────────────────────────
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="裁判書資料", index=True, index_label="序號")
        ws = writer.sheets["裁判書資料"]

        _header_style(ws)
        _data_style(ws, len(rows))
        _set_col_widths(ws, ["序號"] + col_labels)

        # 凍結標題列與序號欄
        ws.freeze_panes = "C2"

        # 自動篩選
        ws.auto_filter.ref = ws.dimensions

        # ── 統計摘要頁 ────────────────────────────────────────────
        wb  = writer.book
        wss = wb.create_sheet("統計摘要")

        title_font = Font(bold=True, size=14, name="微軟正黑體")
        head_font  = Font(bold=True, size=11)

        wss["A1"] = "裁判書匯出統計"
        wss["A1"].font = title_font
        wss.row_dimensions[1].height = 28

        wss["A3"] = "匯出筆數";  wss["A3"].font = head_font;  wss["B3"] = len(rows)
        wss["A4"] = "匯出時間";  wss["A4"].font = head_font
        wss["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # 法院分布
        wss["A6"] = "法院分布";  wss["A6"].font = head_font
        court_count: Dict[str, int] = {}
        for r in rows:
            k = r.get("court", "") or "（未知）"
            court_count[k] = court_count.get(k, 0) + 1
        for i, (c, cnt) in enumerate(sorted(court_count.items(), key=lambda x: -x[1]), 7):
            wss[f"A{i}"] = c
            wss[f"B{i}"] = cnt

        # 案件類型分布
        start_row = 7 + len(court_count) + 2
        wss[f"A{start_row}"] = "案件類型分布"
        wss[f"A{start_row}"].font = head_font
        type_count: Dict[str, int] = {}
        for r in rows:
            k = r.get("case_type", "") or "（未知）"
            type_count[k] = type_count.get(k, 0) + 1
        for i, (t, cnt) in enumerate(sorted(type_count.items(), key=lambda x: -x[1]), start_row + 1):
            wss[f"A{i}"] = t
            wss[f"B{i}"] = cnt

        wss.column_dimensions["A"].width = 30
        wss.column_dimensions["B"].width = 12

    logger.info("Exported %d records → %s", len(rows), output_path)
    return True


# ─── CLI ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="匯出裁判書資料至 Excel")
    ap.add_argument("-n",  "--num",        type=int, default=100,
                    help="匯出筆數，0 表示全部 (預設: 100)")
    ap.add_argument("-o",  "--output",     default="",
                    help="輸出檔名 (預設: judgments_YYYYMMDD_HHMMSS.xlsx)")
    ap.add_argument("-k",  "--keyword",    default="",
                    help="篩選關鍵字（同時比對 keyword / 裁判字號 / 主文）")
    ap.add_argument("-c",  "--court",      default="",
                    help="篩選法院（部分比對）")
    ap.add_argument("--start-date",        default="",
                    help="裁判日期起")
    ap.add_argument("--end-date",          default="",
                    help="裁判日期迄")
    ap.add_argument("--offset", type=int,  default=0,
                    help="略過前 N 筆")
    ap.add_argument("--full-text",         action="store_true",
                    help="包含全文欄位（檔案較大）")
    args = ap.parse_args()

    limit  = args.num if args.num > 0 else None
    output = args.output or f"judgments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    data = fetch_judgments(
        limit=limit,
        offset=args.offset,
        keyword=args.keyword or None,
        court=args.court or None,
        start_date=args.start_date or None,
        end_date=args.end_date or None,
    )

    if not data:
        print("資料庫中無符合條件的資料。請先執行 crawler.py 和 parser.py。")
    else:
        ok = export_to_excel(data, output, include_full_text=args.full_text)
        if ok:
            print(f"\n完成！共匯出 {len(data)} 筆 → {output}")
        else:
            print("匯出失敗。")
